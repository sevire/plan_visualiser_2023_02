"""
Module which reads in plan data from the file in any supported format, and
then parses the data to extract information for each activity.
"""
import re
from abc import ABC, abstractmethod
from datetime import datetime, date
from pathlib import Path
from typing import List, Dict
from plan_visual_django.exceptions import (
    SuppliedPlanIncompleteError,
    ExcelPlanSheetNotFound
)
from plan_visual_django.models import Plan
import openpyxl as openpyxl
import logging
from plan_visual_django.services.plan_file_utilities.plan_field import PlanFieldEnum, FileType, \
    PlanInputFieldSpecification, PlanInputField, PlanFieldInputSourceEnum

logger = logging.getLogger(__name__)

# Utility functions for parsing into every data type required by the app from every plan format
# required to be supported.

# Define conversion functions for each input/output type which needs to be supported


def convert_pass_through(x):
    return x


def convert_str_or_int_to_str(string_or_int) -> str:
    """
    Mostly used in case when we are really expecting a string but sometimes, because of the way in which
    exports from Excel are often generated, a field with digits in will come over as an integer even if
    it was intended as a string.

    If it is an integer, we want to create as simple a string representation of the integer as possible,
    to avoid potential anomalies if we do any zero padding or other formatting as the value will need to
    be re-created every time (for sticky UID for example).

    :param string_or_int:
    :return:
    """
    if isinstance(string_or_int, int):
        return str(string_or_int)
    elif isinstance(string_or_int, str):
        return string_or_int
    else:
        raise ValueError(f"String or int expected, but got {type(string_or_int)} for {string_or_int}")


def convert_str_yes_no_to_bool(string_yes_no) -> bool:
    """
    Used initially for MS Project export to decode the Milestone flag which exports to a string with 'Yes' or 'No'

    :param string_or_int:
    :return:
    """
    if string_yes_no == "Yes":
        return True
    elif string_yes_no == "No":
        return False
    else:
        raise ValueError(f"String value 'Yes' or 'No' expected, but got {type(string_yes_no)} for {string_yes_no}")


def convert_string_int(string) -> int:
    return int(string.strip())


def convert_string_nnd_int(string) -> int:
    if isinstance(string, str):
        return int(('0' + string.strip())[:-1])
    else:
        raise ValueError(f"Conversion expected string of form nnd but got type {type(string)}")


def convert_string_msp_duration_int(string) -> int:
    """
    Parse a string of the form 'nn day', 'nn days', 'nn hour', 'nn hours' or 'nn week', 'nn weeks' with an optional ? at the end to an
    integer number of days.

    :param string:
    :return:
    """
    if string is None:
        # Special case for now - return 0 if the string is None
        print("Warning: convert_string_msp_duration_int called with None value")
        return 0
    if isinstance(string, str):
        matches = re.match(r"(\d+) (\w+)", string.strip())
        if matches is not None:
            duration = int(matches.group(1))
            if matches.group(2) == 'day' or matches.group(2) == 'days':
                return duration
            elif matches.group(2) == 'hour' or matches.group(2) == 'hours' or matches.group(2) == 'hr' or matches.group(2) == 'hrs':
                return duration // 24 + 1
            elif matches.group(2) == 'week' or matches.group(2) == 'weeks':
                return duration * 7
            else:
                raise ValueError(f"Unsupported duration unit {matches.group(2)}")
        else:
            raise ValueError(f"Invalid string {string}")
    else:
        raise ValueError(f"Conversion expected string of form nnd but got type {type(string)}")


def convert_string_float(string) -> float:
    return float(string.strip())


def convert_string_date_dmy_01(date_str) -> date:
    return datetime.strptime(date_str, '%d:%m:%Y').date()


def convert_string_date_dmy_02(date_str) -> date:
    """
    Format: '10 June 2022 08:00'
    :param date_str:
    :return:
    """

    return datetime.strptime(date_str, '%d %B %Y %H:%M').date()


def convert_int_string(int_val) -> str:
    return str(int_val)


def convert_int_float(int_val, float):
    return float(int_val)


def convert_float_string(float_val) -> str:
    return str(float_val)


def convert_float_int(float_val) -> int:
    return int(float_val)


def convert_float_str(float_val) -> str:
    return str(float_val)


    # Drives conversion of input types/encoding to appropriate output types.  Utility to allow input fields to be encoded
    # in different ways (e.g. dates can be encoded in a string in many ways)
convert_dispatch_table = {
    'STR': {
        'STR': convert_pass_through,
        'INT': convert_string_int,
    },
    "STR_DATE_DMY_01": {
        'DATE': convert_string_date_dmy_01
    },
    "STR_DATE_DMY_02": {
        'DATE': convert_string_date_dmy_02
    },
    'STR_OR_INT': {
        'STR': convert_str_or_int_to_str,
    },
    'STR_MSTONE_YES_NO': {
        'BOOL': convert_str_yes_no_to_bool,
    },
    'STR_nnd': {  # Typically used to decode a duration encoded as a number of days e.g. '345d'
        'INT': convert_string_nnd_int,
    },
    'STR_duration_msp': {  # Typically used to decode a duration encoded as a number of days e.g. '345d'
        'INT': convert_string_msp_duration_int,
    },
    'INT': {
        'STR': convert_int_string,
        'INT': convert_pass_through,
    },
    'FLOAT': {
        'INT': convert_float_int,
        'STR': convert_float_str,
    },
    'DATE': {
        'DATE': convert_pass_through,
    }
}


def convert_dispatch(input_type: str, output_type: str, input_value: any) -> any:
    """
    Select the right conversion function to convert this input value to an appropriate output value.

    :param input_type:
    :param output_type:
    :param input_value:
    :return:
    """
    convert_function = convert_dispatch_table[input_type][output_type]
    converted_value = convert_function(input_value)

    return converted_value


class PlanParser():
    """
    Takes data from a plan which has already been read in into rows of data, with one row for each activity within the
    plan, and parses the columns from each activity according to the defined parsing rules, into the correct
    type and value for each field required in order to manage the plan.
    """

    def __init__(self, plan_field_mapping: dict[PlanFieldEnum, PlanInputFieldSpecification]):
        """
        plan_data will be a List of Dictionaries, with one dictionary entry for each colum in the plan file which has
        been read.

        :param plan_data:
        """
        self.plan_field_mapping = plan_field_mapping

    def parse(self, data: List[Dict], headings: List) -> List[Dict]:
        """
        Take raw data from the plan and parse it into the correct type and value for each field required in order to
        construct the plan in the database.

        :param data:
        :param headings:
        :return:
        """
        # Get actual mapping of input fields to the target plan fields from input file
        supplied_field_mapping: List[(PlanFieldEnum, PlanInputFieldSpecification)] = self.validate_input_fields(headings)

        parsed_data = []

        for plan_record in data:
            parsed_data_record = {}
            ignore_record = False

            # We now have a definitive set of input fields mapped to the plan fields we need to store for the plan.
            # Even so, we implement a touch of belt and braces here, checking that each field is in the received fields
            # before processing them.
            # ToDo: Re-visit whether need to check input field is in headings as we have already confirmed this by now.

            plan_field: PlanFieldEnum
            input_field: PlanInputFieldSpecification
            for plan_field, input_field in supplied_field_mapping:
                # Check whether this field is included in the mapping, and ignore field if not.
                if input_field.input_field_name not in headings:
                    parsed_data_record[plan_field.value.field_name] = "(n/a)"
                else:
                    # If we've had an error on this record then we will ignore remaining fields and not include the record
                    # in the plan.
                    if not ignore_record:
                        # Should get exactly one match unless the field is optional and not included in the mapping
                        mapped_field_column_raw_value = plan_record[input_field.input_field_name].get_input_data(input_field.input_field_source)
                        try:
                            field_parsed_value = convert_dispatch(input_field.input_field_type.code, plan_field.value.field_type.value, mapped_field_column_raw_value)
                        except (ValueError, TypeError) as e:
                            print(f"Error parsing record {plan_record}, ignoring record)")
                            ignore_record = True
                        else:
                            parsed_data_record[plan_field.value.field_name.value] = field_parsed_value
            if not ignore_record:
                parsed_data.append(parsed_data_record)

        return parsed_data

    def validate_input_fields(self, headings):
        """
        Check that the fields that are supplied in the input file include all compulsory fields needed for the plan.

        :param headings:
        :return:
        """
        supplied_fields = [plan_field for plan_field, input_field in self.plan_field_mapping.items() if
                           input_field.input_field_name in headings]
        compulsory_mapped_fields = [plan_field for plan_field, input_field in self.plan_field_mapping.items() if plan_field.value.required_flag is True]
        missing_compulsory_mapped_fields = [field for field in compulsory_mapped_fields if field not in supplied_fields]
        if len(missing_compulsory_mapped_fields) > 0:
            logger.error(f"Missing compulsory fields...")
            user_error_message = ""
            prefix = ""
            for missing_field in missing_compulsory_mapped_fields:
                left_arrow = '\u2190'
                missing_field_log_message = f"{missing_field.value.field_name.value} {left_arrow} {self.plan_field_mapping[missing_field].input_field_name}"
                logger.error(missing_field_log_message)
                user_error_message += prefix + missing_field_log_message
                prefix = ", "
            user_error_message = "Some expected fields not present in file: " + user_error_message
            raise SuppliedPlanIncompleteError(user_error_message)
        return [(plan_field, self.plan_field_mapping[plan_field]) for plan_field in supplied_fields]


class PlanFileReader(ABC):
    """
    Base abstract class which defines key operations for a given technical file format.
    This class reads in the data into a List of Dictionaries with:
    - One List entry per activity from the plan.
    - One Dictionary entry per column from the input format.
    - Where possible and appropriate, un-required columnns will be discarded.

    It then parses the data with the supplied parser which will then:
    - Extract required fields using the appropriate columns for the supplied file
    - Parse the fields into the correct type/value.

    The types of file that we will want to support (eventually) will include:
    - Excel (a different reader for each Excel version that needs to be supported).
    - CSV
    - MS Project (native (XML) - not Excel import)


    The generic flow which is defined to allow appropriate hooks
    """
    file_type_name = "(Undefined)"

    def __init__(self):
        self.parser = PlanParser

        self.pre_processing()
        self.post_processing()

    def pre_processing(self):
        """
        Allows any one off logic to be carried out before the reading of the file commences, which is
        dependent upon file type
        :return:
        """
        pass

    @abstractmethod
    def read(self, pathname: str) -> (List[Dict],List):
        """
        Override with code to read records from specific format
        :param pathname:
        :return: List of records and list of heading names of input file
        """
        pass

    def post_processing(self):
        pass


class ExcelXLSFileReader(PlanFileReader):
    def __init__(self, sheet_name: str=None):
        """
        If sheet name is None then use name of only sheet which exists.  If is more than one sheet
        it's an error.

        :param sheet_name:
        """

        super().__init__()
        self.sheet_name = sheet_name

    def get_sheet_name(self, sheet_list: [str], file_name: str, file_type_name: str):
        """
        Uses some fuzzy(ish) logic to work out the sheet name to use for the plan data.

        Logic is:
        - If self.sheet_name has been set, use that.
        - If there is only one sheet, use that.
        - If there is more than one sheet, then test for
          - A sheet with the same name as the workbook (this is default for SmartSheet)
          - A sheet with the name "Task_Data" (this is default for MS Project)
        - If none matches, then abort with fatal error (don't just try to read a random sheet!)

        :param file_type_name: File type of the file, which provides hints as to which sheet the plan data is in.
        :param file_name:
        :param sheet_list:
        :param workbook_object:
        :return:
        """
        if self.sheet_name is not None and self.sheet_name != "":
            sheet_name = self.sheet_name
            logger.debug(f"Sheet name set outside this method, using that value, sheet name = {sheet_name}")
            return sheet_name
        elif len(sheet_list) == 1:
            sheet_name = sheet_list[0]
            logger.debug(f"Only one sheet in the plan, using that, sheet name = {sheet_name}")
            return sheet_name
        elif file_type_name == "excel-02-smartsheet-export-01" and file_name in sheet_list:
            sheet_name = file_name
            logger.debug(f"Smartsheet, choosing sheet name = file name, sheet name = {sheet_name}")
            return sheet_name
        elif file_type_name == "excel-01-msp-export-default-01" and "Task_Data" in sheet_list:
            sheet_name = "Task_Data"
            logger.debug(f"MS Project, choosing sheet name = {sheet_name}")
            return sheet_name
        else:
            message = f"Unable to determine which sheet plan is in within file {file_name}, aborting"
            logger.error(message)
            raise ExcelPlanSheetNotFound(message)

    def read(self, plan: Plan) -> (List[Dict], List):
        """
        Reads Excel file to extract data for activities in the plan.

        Key features:
        - Each row represents one activity.
        :return:
        """
        skip_rows = 0

        wb_obj = openpyxl.load_workbook(plan.file.file, data_only=True)
        name = Path(plan.file_name).stem

        sheet_name = self.get_sheet_name(wb_obj.sheetnames, name, plan.file_type_name)
        sheet = wb_obj[sheet_name]
        start_row = 1 + skip_rows

        headings = self.get_headers(sheet, start_row)
        table = {heading: [] for heading in headings}

        num_blank_rows = 0
        max_blank_rows = 100
        read_row_num = start_row
        while num_blank_rows <= max_blank_rows:
            row_confirmed = self.read_row(read_row_num, sheet, headings, table)
            if not row_confirmed:
                num_blank_rows += 1
            read_row_num += 1

        iterable_by_row = self.iterrows(table)
        return iterable_by_row, headings

    def parse(self, raw_data, raw_data_headers, plan_field_mapping):
        parser = self.parser(plan_field_mapping=plan_field_mapping)
        return parser.parse(raw_data, raw_data_headers)

    @staticmethod
    def read_row(table_row_num, sheet, headings, table, skiprows=0):
        maybe_row = {}
        row_confirmed = False
        for col, heading in enumerate(headings):
            cell = sheet.cell(table_row_num + skiprows + 1, col + 1)

            # Now capture both value and indent level as either could be needed in parsing.
            cell_value = cell.value
            cell_indent = cell.alignment.indent

            maybe_row[heading] = PlanInputField(
                value=cell_value,
                indent=cell_indent
            )
            if cell_value is not None:
                row_confirmed = True
        if row_confirmed:
            for heading in headings:
                table[heading].append(maybe_row[heading])
        return row_confirmed

    @staticmethod
    def get_headers(sheet, start_row):
        headings = []  # Need to store headings in column order so list not dict
        blank = False
        row = start_row
        column = 1
        while not blank:
            maybe_heading = sheet.cell(row, column).value
            if maybe_heading is not None:
                headings.append(maybe_heading)
                column += 1
            else:
                blank = True
        return headings

    @staticmethod
    def iterrows(excel_table):
        # Work out num rows by checking any of the column arrays
        num_rows = len(next(iter(excel_table.values())))
        return [{heading: excel_table[heading][row] for heading in excel_table} for row in range(num_rows)]
